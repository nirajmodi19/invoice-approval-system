import os
from openpyxl import Workbook
import bcrypt

EXCEL_FILE = 'data.xlsx'

def hash_password(password: str) -> str:
    # Hash a password and return it as a string
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed.decode('utf-8')

def create_excel():
    if os.path.exists(EXCEL_FILE):
        print(f"{EXCEL_FILE} already exists. Skipping creation.")
        return

    wb = Workbook()

    # Create Users sheet
    ws_users = wb.active
    ws_users.title = 'Users'
    ws_users.append(['Username', 'Password', 'Role'])

    # Default users with hashed passwords
    users = [
        ('manager', hash_password('manager123'), 'Manager'),
        ('accounts', hash_password('accounts123'), 'Accounts'),
        ('ceo', hash_password('ceo123'), 'CEO'),
    ]
    for user in users:
        ws_users.append(user)

    # Create Invoices sheet
    ws_invoices = wb.create_sheet(title='Invoices')
    ws_invoices.append(['Invoice ID', 'Description', 'Amount', 'Raised By', 'Status', 'Approved By (Accounts)', 'Approved By (CEO)'])

    # Create Approvals sheet
    ws_approvals = wb.create_sheet(title='Approvals')
    ws_approvals.append(['Invoice ID', 'Approver', 'Status', 'Timestamp'])

    wb.save(EXCEL_FILE)
    print(f"{EXCEL_FILE} created successfully with default users.")

if __name__ == '__main__':
    create_excel()
