import openpyxl
from datetime import datetime

EXCEL_FILE = 'invoice_system.xlsx'

# Initialize Excel file if not present
def initialize_excel():
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Invoices'
        sheet.append(['ID', 'Description', 'Amount', 'Raised By', 'Status', 'Created At', 'File Path'])
        workbook.create_sheet('Users')
        user_sheet = workbook['Users']
        user_sheet.append(['Username', 'Password', 'Role'])
        # Add sample users
        user_sheet.append(['manager', 'manager123', 'Manager'])
        user_sheet.append(['accounts', 'accounts123', 'Accounts'])
        user_sheet.append(['ceo', 'ceo123', 'CEO'])
        
        # Approval Logs Sheet
        workbook.create_sheet('ApprovalLogs')
        log_sheet = workbook['ApprovalLogs']
        log_sheet.append(['Invoice ID', 'Role', 'Action', 'Timestamp'])
        workbook.save(EXCEL_FILE)
    finally:
        workbook.close()

# Function to validate user login
def validate_user(username, password):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['Users']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            return True
    return False

# Function to get user details
def get_user(username):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['Users']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return {'username': row[0], 'role': row[2]}
    return None

# Function to raise a new invoice
def add_invoice(description, amount, raised_by, file_path=None):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['Invoices']
    next_id = sheet.max_row if sheet.max_row > 1 else 1
    sheet.append([
        next_id, description, amount, raised_by, 
        'Pending', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 
        file_path
    ])
    workbook.save(EXCEL_FILE)

# Function to get invoices based on user role
def get_invoices_for_role(role, username):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['Invoices']
    invoices = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        invoice = {
            'id': row[0],
            'description': row[1],
            'amount': row[2],
            'raised_by': row[3],
            'status': row[4],
            'created_at': row[5],
            'file_path': row[6]
        }

        if role == 'Manager' and invoice['raised_by'] == username:
            invoices.append(invoice)
        elif role == 'Accounts' and invoice['status'] == 'Pending':
            invoices.append(invoice)
        elif role == 'CEO' and invoice['status'] == 'Approved by Accounts':
            invoices.append(invoice)
    
    return invoices

# Function to update invoice status
def update_invoice_status(invoice_id, status, role):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['Invoices']
    
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == invoice_id:
            row[4].value = status
            break
    
    # Add approval log
    add_approval_record(invoice_id, role, status)
    
    workbook.save(EXCEL_FILE)

# Function to add approval record (logs)
def add_approval_record(invoice_id, role, status):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    log_sheet = workbook['ApprovalLogs']
    log_sheet.append([invoice_id, role, status, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    workbook.save(EXCEL_FILE)

# Function to get approval logs for a specific invoice
def get_approval_logs(invoice_id):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    log_sheet = workbook['ApprovalLogs']
    logs = []

    for row in log_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == invoice_id:
            logs.append({'role': row[1], 'action': row[2], 'timestamp': row[3]})
    
    return logs

# Initialize Excel on the first run
initialize_excel()
